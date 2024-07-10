from flask import Flask, jsonify, send_file, request
import pandas as pd
from sklearn.preprocessing import MinMaxScaler
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import boto3
from botocore.exceptions import ClientError
import matplotlib
matplotlib.use('Agg')  # Set the backend
from datetime import datetime, timedelta
import requests
import os
from dotenv import load_dotenv
import tensorflow as tf  # Import for early stopping

load_dotenv()

app = Flask(__name__)

# --- Data ---
def fetch_data_from_api(collection, branchId):
    url = "https://ap-south-1.aws.data.mongodb-api.com/app/data-kbpgj/endpoint/data/v1/action/find"
    headers = {
        "Content-Type": "application/json",
        "api-key": os.getenv('MONGODB_API_KEY')
    }
    payload = {
        "collection": collection,
        "database": "flipper",
        "dataSource": "Cluster0",
        "filter": {"branchId": branchId}
    }

    try:
        response = requests.post(url, json=payload, headers=headers)
        response.raise_for_status()
        return response.json().get('documents', [])
    except requests.RequestException as e:
        print(f"Error fetching {collection} data from API: {e}")
        return []

def fetch_products(branchId):
    products_data = fetch_data_from_api("Variant", branchId)
    return {
        doc.get('id'): {
            "id": doc.get('id'),
            "name": doc.get('productName'),
            "cost": doc.get('supplyPrice', 0),
            "price": doc.get('retailPrice', 0)
        }
        for doc in products_data
    }

def fetch_stock(branchId):
    stock_data = fetch_data_from_api("Stock", branchId)
    return {
        doc.get('variantId'): doc.get('currentStock', 0)
        for doc in stock_data
    }

def fetch_sales(branchId):
    sales_data = fetch_data_from_api("TransactionItem", branchId)
    processed_sales = {}
    
    for item in sales_data:
        date = pd.to_datetime(item.get('createdAt')).date()
        variant_id = item.get('variantId')
        quantity = item.get('qty', 0)
        
        if variant_id not in processed_sales:
            processed_sales[variant_id] = {'date': [], 'sales': []}
        
        if date in processed_sales[variant_id]['date']:
            index = processed_sales[variant_id]['date'].index(date)
            processed_sales[variant_id]['sales'][index] += quantity
        else:
            processed_sales[variant_id]['date'].append(date)
            processed_sales[variant_id]['sales'].append(quantity)
    
    return processed_sales

# --- Functions ---

def prepare_data(df, lookback=1):
    X, y = [], []
    for i in range(lookback, len(df)):
        X.append(df['sales'].values[i-lookback:i])
        y.append(df['sales'].values[i])
    return np.array(X), np.array(y)

def train_model(product_id, sales_data, lookback=1, batch_size=32, validation_split=0.1):
    """
    Trains an LSTM model for sales prediction for a given product.

    Args:
        product_id: The ID of the product.
        sales_data: A dictionary containing sales data for the product.
        lookback: The number of past time steps to use as input for the LSTM model.
        batch_size: The batch size for training.
        validation_split: The proportion of data to use for validation.

    Returns:
        The trained LSTM model.
    """

    if product_id not in sales_data or len(sales_data[product_id]['date']) < 2:
        print(f"Insufficient sales data for product {product_id}")
        return None

    # Create a Pandas DataFrame from the sales data
    df = pd.DataFrame(sales_data[product_id])
    df['date'] = pd.to_datetime(df['date'])
    df = df.set_index('date')
    df = df.sort_index()
    df = df.resample('D').sum().fillna(0)  # Resample to daily data, filling missing days with 0

    # Data Preprocessing
    scaler = MinMaxScaler(feature_range=(0, 1))
    df['sales'] = scaler.fit_transform(df['sales'].values.reshape(-1, 1))

    # Prepare the data for the LSTM model
    X, y = prepare_data(df, lookback)
    X = X.reshape(X.shape[0], X.shape[1], 1)  # Reshape for LSTM input

    # Create the LSTM model
    model = Sequential()
    model.add(LSTM(50, return_sequences=True, input_shape=(X.shape[1], 1)))
    model.add(LSTM(50, return_sequences=False))
    model.add(Dense(25))
    model.add(Dense(1))
    model.compile(optimizer='adam', loss='mean_squared_error')

    # Early Stopping to prevent overfitting
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=10, restore_best_weights=True)

    # Train the model with validation split (or use validation_data)
    if len(X) > 1:  # Check if we have enough data for validation
        model.fit(X, y, epochs=100, batch_size=batch_size, validation_split=validation_split, callbacks=[early_stopping])
    else:
        print(f"Skipping training for product {product_id} due to insufficient data.")
        return None 

    # Save the trained model
    model.save(f'trained_models/{product_id}_model.h5')

    return model

def prepare_data(df, lookback=1):
    """
    Prepares the data for the LSTM model.

    Args:
        df: The Pandas DataFrame containing the sales data.
        lookback: The number of past time steps to use as input.

    Returns:
        X: A NumPy array of the input data.
        y: A NumPy array of the target data.
    """
    X, y = [], []
    for i in range(lookback, len(df)):
        X.append(df['sales'].values[i-lookback:i])
        y.append(df['sales'].values[i])
    return np.array(X), np.array(y)

def predict_sales(product_id, days_to_predict=7):
    try:
        model = tf.keras.models.load_model(f'trained_models/{product_id}_model.h5')  # Load the model
    except FileNotFoundError:
        print(f"No trained model found for product {product_id}, training a new one.")
        # Train the model for a specific product
        model = train_model(product_id, sales_data, validation_split=0.1)
        if not model:
            return []  # Return empty list if training failed

    if product_id not in sales_data or len(sales_data[product_id]['date']) < 2:
        print(f"Insufficient sales data for product {product_id}")
        return []

    df = pd.DataFrame(sales_data[product_id])
    df['date'] = pd.to_datetime(df['date'])
    df = df.set_index('date')
    df = df.sort_index()
    df = df.resample('D').sum().fillna(0)  # Resample to daily data, filling missing days with 0

    # Data Preprocessing (Apply scaling here)
    scaler = MinMaxScaler(feature_range=(0, 1))
    df['sales'] = scaler.fit_transform(df['sales'].values.reshape(-1, 1))

    last_day_sales = df['sales'].values[-1].reshape(1, 1, 1)
    predictions = []

    for _ in range(days_to_predict):
        next_day_prediction = model.predict(last_day_sales)
        predictions.append(next_day_prediction[0, 0])
        last_day_sales = next_day_prediction.reshape(1, 1, 1)

    predictions = scaler.inverse_transform(np.array(predictions).reshape(-1, 1))
    
    # Apply a smoothing factor to prevent extreme fluctuations
    smoothing_factor = 0.8
    smoothed_predictions = [predictions[0][0]]
    for i in range(1, len(predictions)):
        smoothed_predictions.append(smoothing_factor * predictions[i][0] + (1 - smoothing_factor) * smoothed_predictions[i-1])
    
    return np.array(smoothed_predictions)

def generate_excel_file():
    if not sales_data:
        raise ValueError("No sales data available")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Predictions"
    headers = ['Date', 'Product Name', 'Predicted Sales', 'Current Stock', 'Stock Replenishment', 'Stock Value', 'Gross Profit', 'Cost of Goods Sold', 'Net Profit']
    ws.append(headers)

    # Track added products to avoid duplicates
    added_products = set()

    # Deduplicate products based on product names
    unique_products = {}
    for product_id, product in products.items():
        product_name = product['name']
        if product_name not in unique_products:
            unique_products[product_id] = product

    for product_id, product in unique_products.items():
        product_name = product['name']
        if product_name in added_products:  # Check if product name is already added
            continue

        predicted_sales = predict_sales(product_id)

        if len(predicted_sales) == 0:
            continue
        
        current_stock = stock.get(product_id, 0)  # Get the current stock for this product
        current_date = datetime.now().date()
        
        output_data = []
        
        for i, sales_prediction in enumerate(predicted_sales):
            date = current_date + timedelta(days=i)
            target_stock_level = max(7 * sales_prediction, current_stock * 0.9)  # Keep at least 90% of current stock
            replenishment_needed = max(0, target_stock_level - current_stock)
            stock_value = current_stock * product['price']
            
            # Calculate profit using actual data
            sales_on_this_date = sales_data[product_id]['sales'][i] if i < len(sales_data[product_id]['sales']) else 0
            gross_profit = sales_on_this_date * product['price'] - sales_on_this_date * product['cost']
            cost_of_goods_sold = sales_on_this_date * product['cost']
            net_profit = gross_profit 
            
            output_data.append([
                date,
                product_name,
                round(sales_prediction, 2),
                round(current_stock, 2),
                round(replenishment_needed, 2),
                round(stock_value, 2),
                round(gross_profit, 2),  # Add gross profit
                round(cost_of_goods_sold, 2),  # Add cost of goods sold
                round(net_profit, 2)  # Add net profit
            ])
            
            # Update current_stock for the next day
            current_stock = max(0, current_stock - sales_prediction + replenishment_needed)
        
        for row in output_data:
            ws.append(row)

        # Add product name to the list of added products
        added_products.add(product_name)

    # Create chart
    chart = LineChart()
    chart.title = "Sales Predictions and Stock Levels"
    chart.style = 10
    chart.y_axis.title = 'Quantity'
    chart.x_axis.title = 'Date'
    
    data = Reference(ws, min_col=3, min_row=1, max_col=5, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    chart.width = 20
    chart.height = 12
    
    ws.add_chart(chart, "H2")
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def generate_file_name():
    current_date = datetime.now().strftime("%Y%m%d")
    return f"REPORT_{current_date}.xlsx".upper()

@app.route('/generate_s3_url', methods=['GET'])
def generate_s3_url():
    branchId = request.args.get('branchId')
    if not branchId:
        return jsonify({
            "status": "error",
            "message": "branchId is required"
        }), 400

    # Fetch data using the provided branchId
    global products, stock, sales_data
    products = fetch_products(int(branchId))
    stock = fetch_stock(int(branchId))
    sales_data = fetch_sales(int(branchId))

    if not sales_data:
        return jsonify({
            "status": "error",
            "message": "No sales data available for the given branchId"
        }), 404

    bucket_name = os.getenv('BUCKET_NAME')
    file_name = generate_file_name()
    
    s3_client = boto3.client('s3')
    
    try:
        excel_buffer = generate_excel_file()
    except ValueError as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 404
    
    # Save file locally
    with open(file_name, 'wb') as f:
        f.write(excel_buffer.getvalue())

    # Upload report to S3
    s3_path = f"public/reports-{branchId}/{file_name}" 
    try:
        s3_client.upload_file(file_name, bucket_name, s3_path)
        print(f"File uploaded successfully to {bucket_name}/{s3_path}")

        # Upload the 'trained_models' folder to S3
        upload_folder_to_s3(s3_client, bucket_name, 'trained_models', f"public/trained_models")

        # Generate presigned URL for the report
        try:
            response = s3_client.generate_presigned_url('get_object',
                                                        Params={'Bucket': bucket_name,
                                                                'Key': s3_path},
                                                        ExpiresIn=3600)
            return jsonify({
                "status": "success",
                "fileName": file_name,
                "download_url": response
            }), 200
        except ClientError as e:
            print(f"Error generating presigned URL: {str(e)}")
            return jsonify({
                "status": "error",
                "message": "Failed to generate downloadable URL"
            }), 500
    except Exception as e:
        print(f"Error uploading file to S3: {str(e)}")
        return jsonify({
            "status": "error",
            "message": "Failed to upload file to S3"
        }), 500



def upload_folder_to_s3(s3_client, bucket_name, local_folder_path, s3_prefix):
    """Uploads a local folder to S3.

    Args:
        s3_client: A boto3 S3 client.
        bucket_name: The name of the S3 bucket.
        local_folder_path: The path to the local folder.
        s3_prefix: The S3 prefix to use for the folder.
    """
    for root, _, files in os.walk(local_folder_path):
        for file in files:
            local_file_path = os.path.join(root, file)
            s3_key = os.path.relpath(local_file_path, local_folder_path)
            s3_key = os.path.join(s3_prefix, s3_key)  # Construct the S3 key

            try:
                s3_client.upload_file(local_file_path, bucket_name, s3_key)
                print(f"Uploaded {local_file_path} to {bucket_name}/{s3_key}")
            except ClientError as e:
                print(f"Error uploading {local_file_path}: {str(e)}")

@app.route('/download_excel', methods=['GET'])
def download_excel():
    excel_buffer = generate_excel_file()
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name='sales_predictions.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True) 